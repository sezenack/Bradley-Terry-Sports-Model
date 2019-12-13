import xlrd
import openpyxl

# Dictionary of every team's rating; key = team name, value = rating
ratings = dict()
# Dictionary of every team's record; key = team name, value = tuple of wins and losses
record = dict()
# Dictionary of every team's expected wins; key = team name, value = expected wins
expected = dict()
# List of tuples which are the games; each tuple is (winner, loser)
games = []
# Open the spreadsheet and assign the first 2 sheets
# Replace with your xlsx file name
wb = xlrd.open_workbook('Bradley-Terry Spreadsheet NCAA Hockey.xlsx') 
firstsheet = wb.sheet_by_index(0)
secondsheet = wb.sheet_by_index(1)
# Dictionary of every team's new rating calculated; key = team name, value = new rating
new_ratings = dict()
# Dictionary of every team's info; key = team name, value = list of [win %, win ratio, strength of schedule]
info = dict()
# Value to compare every rating to and then do recursion
DELTA = 0.0001
# Main boolean used with while loop for recursion based on flag
done = False
# If all values are within DELTA (i.e. should recursion finish)
flag = True

# Read in all the teams and info
r = 1
while r < firstsheet.nrows:
    # Read in team and their record and add to records dict
    team = firstsheet.cell_value(r, 0)
    record[team] = (firstsheet.cell_value(r, 1) + 0.5 * firstsheet.cell_value(r, 3), \
                    firstsheet.cell_value(r, 2) + 0.5 * firstsheet.cell_value(r, 3))
    # Always start with 100 as every team's rating (using iteration to solve the recursive problem)
    ratings[team] = 100
    # If team is undefeated, we'll get a divide by 0 error, so we set the ratio to 25
    # Note: this formula is most effective when there are no undefeated or winless teams
    # and a chain of wins (or ties) can be made from every team to any other team
    if record[team][1] == 0:
        ratio = 25
    # If the team is winless, everything would automatically be 0, so set the ratio to 1/25
    elif record[team][0] == 0:
        ratio = 1/25
    # Win ratio = wins / losses
    else:
        ratio = record[team][0] / record[team][1]
    # Add the team's win %, win ratio and sos is intialized to 0 (to be calculated later)
    teaminfo = [record[team][0] / (record[team][0] + record[team][1]), ratio, 0]
    # Add the list of team info to the info dict
    info[team] = teaminfo
    # Increment counter for reading the spreadsheet
    r += 1

# Read in every game with the teams
g = 1
while g < secondsheet.nrows:
    # Add a tuple (winner, loser) to the list of games
    games.append((secondsheet.cell_value(g, 0), secondsheet.cell_value(g, 2)))
    # Increment counter for reading the spreadsheet
    g += 1

# Recursion for getting accurate ratings
while not done:
    # Initialize the flag to True each time
    flag = True
    
    # Clear expected wins each iteration
    expected.clear()
    # For every game, calculate:
    for game in games:
        # Weighting factor (1 divided by the sum of the ratings of the 2 teams)
        wf = 1 / (ratings[game[0]] + ratings[game[1]])
        # Check to see if each team is in expected wins dictionary
        if game[0] in expected:
            # Multiply team's rating by weighting factor and add to sum
            expected[game[0]] += (ratings[game[0]] * wf)
        else:
            # Multiply team's rating by weighting factor and initialize as expected wins
            expected[game[0]] = (ratings[game[0]] * wf)
        if game[1] in expected:
            # Multiply team's rating by weighting factor and add to sum
            expected[game[1]] += (ratings[game[1]] * wf)
        else:
            # Multiply team's rating by weighting factor and initialize as expected wins
            expected[game[1]] = (ratings[game[1]] * wf)
    
    # For every team, calculate:
    for key in ratings:
        # New rating for the team equals the team's wins divided by expected wins multiplied by the old rating
        new_ratings[key] = (record[key][0] / expected[key]) * ratings[key]
        # Update the SOS for the team
        info[key][2] = new_ratings[key] / info[key][1]
        # If the difference between old rating and new rating <= DELTA
        # If flag is true, that means so far every team's new rating has been within DELTA
        # (since flag is initialized to true)
        if abs(ratings[key] - new_ratings[key]) <= DELTA and abs(record[key][0] - expected[key]) <= DELTA and flag:
            done = True
        # If the difference is greater than DELTA, we must continue the recursion
        # If flag is false, one team has already failed the DELTA test and we must continue the recursion
        else:
            flag = False
            done = False
    # After going through all teams, update ratings
    ratings = new_ratings

# Scale the ratings to an average of 100
for i in range(10):
    scale_wins = 0
    for key in ratings:
        scale_wins += 100 / (100 + ratings[key])
    scale = scale_wins / 30
    
    # Adjust every team's rating and SOS according to scale
    for key in ratings:
        ratings[key] *= scale
        info[key][2] = ratings[key] / info[key][1]

# Sort the teams by their rating
sortedratings = sorted(ratings.items(), key=lambda kv: kv[1], reverse = True)
# Replace with your xlsx file name
wb = openpyxl.load_workbook('Bradley-Terry Spreadsheet NCAA Hockey.xlsx')
wb.create_sheet('Output')
# For outputting the results into the spreadsheet
out = wb.get_sheet_by_name('Output')
for r in range(firstsheet.nrows):
    # Header row
    if r == 0:
        out.cell(row = r + 1, column = 1).value = 'Team'
        out.cell(row = r + 1, column = 2).value = 'Rating'
        out.cell(row = r + 1, column = 3).value = 'Win % Rank'
        out.cell(row = r + 1, column = 4).value = 'Record'
        out.cell(row = r + 1, column = 5).value = 'Win %'
        out.cell(row = r + 1, column = 6).value = 'Win Ratio'
        out.cell(row = r + 1, column = 7).value = 'SOS Rank'
        out.cell(row = r + 1, column = 8).value = 'SOS'
    else:
        # Eight columns
        for col in range(8):
            # Col 0 is the team name, col 1 is the rating
            if col < 2:
                out.cell(row = r + 1, column = col + 1).value = sortedratings[r - 1][col]
            # Formula to calculate the rank of winning %
            elif col == 2:
                out.cell(row = r + 1, column = col + 1).value = "=RANK(E" + str(r + 1) + \
                    ",E2:E" + str(firstsheet.nrows) + ")"
            # Formula to display record
            elif col == 3:
                # Change name to whatever you name your first sheet in the file
                out.cell(row = r + 1, column = col + 1).value = "=VLOOKUP(A" + str(r + 1) \
                    + ",'Teams and Records'!$A$2:$B$" + str(firstsheet.nrows) + \
                    ",2,FALSE)&\"-\"&VLOOKUP(A" + str(r + 1) + ",'Teams and Records'!$A$2:$C$" \
                    + str(firstsheet.nrows) + ",3,FALSE)&\"-\"&VLOOKUP(A" + str(r + 1) \
                    + ",'Teams and Records'!$A$2:$D$" + str(firstsheet.nrows) + ",4,FALSE)"
            # Output win % (col 4) and win ratio (col 5)
            elif col < 6:
                out.cell(row = r + 1, column = col + 1).value = info[sortedratings[r - 1][0]][col - 4]
            # Formula to calculate the SOS rank
            elif col == 6:
                out.cell(row = r + 1, column = col + 1).value = "=RANK(H" + str(r + 1) + \
                    ",H2:H" + str(firstsheet.nrows) + ")"
            # Output SOS (col 7)
            else:
                out.cell(row = r + 1, column = col + 1).value = info[sortedratings[r - 1][0]][col - 5]

# Save the sheet with the output
wb.save('sample.xlsx')
