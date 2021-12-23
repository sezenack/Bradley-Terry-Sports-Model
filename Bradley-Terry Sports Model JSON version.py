import openpyxl
import json

def main():
    # Open json file with utf-8 encoding; insert your own json filename
    with open("NCBCA 2062 Post-CT.json", encoding = 'utf-8-sig') as json_file:  
        data = json.load(json_file)

    # List of tuples which are the games; each tuple is (winner, loser)
    games = []
    # Initalize dictionary of teams mapping their id to their name
    teams = dict()

    # Mega dictionary with all info
    # key = team name
    # value = dictionary with: rating (int), record (list of [wins, losses]), expected wins (int)
    # new rating (int) and info (list of [win %, win ratio, strength of schedule])
    teamstats = dict()
    # Value to compare every rating to and then do recursion
    DELTA = 0.0001
    # Main boolean used with while loop for recursion based on flag
    done = False
    # If all values are within DELTA (i.e. should recursion finish)
    flag = True

    # Extract information that we need from the data
    for t in data["teams"]:
        # teams[id] = name
        teams[t["tid"]] = t["region"]
        # Initialize team dict
        teamstats[t["region"]] = dict()
        # Initialize record to 0-0
        teamstats[t["region"]]["record"] = [0, 0]
        # Always start with 100 as every team's rating (using iteration to solve the recursive problem)
        teamstats[t["region"]]["rating"] = 100

    # Create a Workbook
    wb = openpyxl.Workbook()
    # Create sheet to output games to so that postseason games can be easily added
    out = wb.active
    out.title = "Games"
    # Row counter
    r = 1
    # First row
    out.cell(row = r, column = 1).value = 'Winning Team'
    out.cell(row = r, column = 2).value = 'Winning Score'
    out.cell(row = r, column = 3).value = 'Losing Team'
    out.cell(row = r, column = 4).value = 'Losing Score'

    # Go through all the games and add tuple of (winning team, losing team) to list of games
    # In addition output the games to a sheet so that postseason games can be added with ease
    for g in data["games"]:
        winnerid = g["won"]["tid"]
        loserid = g["lost"]["tid"]
        games.append((teams[winnerid], teams[loserid]))

        # Incrememnt record of each team
        teamstats[teams[winnerid]]["record"][0] += 1
        teamstats[teams[loserid]]["record"][1] += 1

        out.cell(row = r + 1, column = 1).value = teams[winnerid]
        out.cell(row = r + 1, column = 2).value = g["won"]["pts"]
        out.cell(row = r + 1, column = 3).value = teams[loserid]
        out.cell(row = r + 1, column = 4).value = g["lost"]["pts"]
        r += 1

    # Update info after every game is read
    for key in teamstats:
        # If team is undefeated or winless, we'll get a divide by 0 error
        # Note: this formula is most effective when there are no undefeated or winless teams
        # and a chain of wins (or ties) can be made from every team to any other team
        ratio = teamstats[key]["record"][0] / teamstats[key]["record"][1]
        # Add the team's win %, win ratio and sos is intialized to 0 (to be summed later)
        teaminfo = [teamstats[key]["record"][0] / (teamstats[key]["record"][0] + teamstats[key]["record"][1]), ratio, 0]
        # Add the list of team info to the info dict
        teamstats[key]["info"] = teaminfo

    # Recursion for getting accurate ratings
    while not done:
        # Initialize the flag to True each time
        flag = True
        
        # Clear expected wins each iteration
        for t in teamstats:
            teamstats[t]["expected"] = 0

        # For every game, calculate:
        for game in games:
            # Weighting factor (1 divided by the sum of the ratings of the 2 teams)
            wf = 1 / (teamstats[game[0]]["rating"] + teamstats[game[1]]["rating"])
            # Multiply team's rating by weighting factor and add to sum
            teamstats[game[0]]["expected"] += (teamstats[game[0]]["rating"] * wf)
            # Multiply team's rating by weighting factor and add to sum
            teamstats[game[1]]["expected"] += (teamstats[game[1]]["rating"] * wf)
        
        # For every team, calculate:
        for key in teamstats:
            # New rating for the team equals the team's wins divided by expected wins multiplied by the old rating
            teamstats[key]["new_rating"] = (teamstats[key]["record"][0] / teamstats[key]["expected"]) * teamstats[key]["rating"]
            
            # Update the SOS for the team
            teamstats[key]["info"][2] = teamstats[key]["new_rating"] / teamstats[key]["info"][1]
            
            # If the difference between old rating and new rating <= DELTA
            # If flag is true, that means so far every team's new rating has been within DELTA
            # (since flag is initialized to true)
            if abs(teamstats[key]["rating"] - teamstats[key]["new_rating"]) <= DELTA \
            and abs(teamstats[key]["record"][0] - teamstats[key]["expected"]) <= DELTA and flag:
                done = True
            # If the difference is greater than DELTA, we must continue the recursion
            # If flag is false, one team has already failed the DELTA test and we must continue the recursion
            else:
                flag = False
                done = False
        
        # After going through all teams, update ratings
        for t in teamstats:
            teamstats[t]["rating"] = teamstats[t]["new_rating"]

    # Scale the ratings to an average of 100
    for i in range(10):
        scale_wins = 0
        for key in teamstats:
            scale_wins += 100 / (100 + teamstats[key]["rating"])
        # Divide by half the total number of teams
        scale = scale_wins / (len(teamstats) / 2)
        
        # Adjust every team's rating and SOS according to scale
        for key in teamstats:
            teamstats[key]["rating"] *= scale
            teamstats[key]["info"][2] = teamstats[key]["rating"] / teamstats[key]["info"][1]

    # Make ratings dict for output
    ratings = dict()
    for t in teamstats:
        ratings[t] = teamstats[t]["rating"]
    # Sort the teams by their rating
    sortedratings = sorted(ratings.items(), key=lambda kv: kv[1], reverse = True)
    # Create new sheet for output
    wb.create_sheet('Output')
    # For outputting the results into the spreadsheet
    out = wb.get_sheet_by_name('Output')
    for r in range(len(sortedratings) + 1):
        # Header row
        if r == 0:
            out.cell(row = r + 1, column = 1).value = 'Rank'
            out.cell(row = r + 1, column = 2).value = 'Team'
            out.cell(row = r + 1, column = 3).value = 'Rating'
            out.cell(row = r + 1, column = 4).value = 'Win % Rank'
            out.cell(row = r + 1, column = 5).value = 'Record'
            out.cell(row = r + 1, column = 6).value = 'Win %'
            out.cell(row = r + 1, column = 7).value = 'Win Ratio'
            out.cell(row = r + 1, column = 8).value = 'SOS Rank'
            out.cell(row = r + 1, column = 9).value = 'SOS'
        else:
            # Nine columns
            for col in range(9):
                # Col 0 is the rank
                if col == 0:
                    out.cell(row = r + 1, column = col + 1).value = r
                # Col 1 is the team name, col 2 is the rating
                elif col < 3:
                    out.cell(row = r + 1, column = col + 1).value = sortedratings[r - 1][col - 1]
                # Formula to calculate the rank of winning %
                elif col == 3:
                    out.cell(row = r + 1, column = col + 1).value = "=RANK(F" + str(r + 1) + \
                        ",F2:F" + str(len(sortedratings) + 1) + ")"
                # Output record
                elif col == 4:
                    out.cell(row = r + 1, column = col + 1).value = str(teamstats[sortedratings[r - 1][0]]["record"][0]) \
                        + "-" + str(teamstats[sortedratings[r - 1][0]]["record"][1])
                # Output win % (col 5) and win ratio (col 6)
                elif col < 7:
                    out.cell(row = r + 1, column = col + 1).value = teamstats[sortedratings[r - 1][0]]["info"][col - 5]
                # Formula to calculate the SOS rank
                elif col == 7:
                    out.cell(row = r + 1, column = col + 1).value = "=RANK(I" + str(r + 1) + \
                        ",I2:I" + str(len(sortedratings) + 1) + ")"
                # Output SOS (col 8)
                else:
                    out.cell(row = r + 1, column = col + 1).value = teamstats[sortedratings[r - 1][0]]["info"][col - 6]

    # Save the sheet with the output
    wb.save('Bradley-Terry Spreadsheet JSON NCBCA.xlsx')
    return ratings


if __name__ == "__main__":
    main()
    