import openpyxl

# Dictionary of every team's rating; key = team name, value = rating
ratings = dict()
# Dictionary of every team's record; key = team name, value = tuple of wins and losses
record = dict()
# Dictionary of every team's expected wins; key = team name, value = expected wins
expected = dict()
# List of tuples which are the games; each tuple is (winner, loser)
games = []
# Open the spreadsheet and assign the first 2 sheets
# Replace with your xlsx file name
wb = openpyxl.load_workbook('Bradley-Terry Spreadsheet NCAA Hockey.xlsx') 
firstsheet = wb.worksheets[0]
# Dictionary of every team's new rating calculated; key = team name, value = new rating
new_ratings = dict()
# Dictionary of every team's info; key = team name, value = list of [win %, win ratio, strength of schedule]
info = dict()
# Value to compare every rating to and then do recursion
DELTA = 0.0001
# Main boolean used with while loop for recursion based on flag
done = False
# If all values are within DELTA (i.e. should recursion finish)
flag = True

# Read in every game with the teams
g = 2
while g < firstsheet.max_row + 1:
    # Skip exhibition games
    marker = firstsheet.cell(row = g, column = 6).value.strip()
    if marker == 'ex' or marker == 'n3':
        g += 1
        continue
    team1 = firstsheet.cell(row = g, column = 1).value.strip()
    team1score = firstsheet.cell(row = g, column = 2).value
    team2 = firstsheet.cell(row = g, column = 3).value.strip()
    team2score = firstsheet.cell(row = g, column = 4).value
    regulation = firstsheet.cell(row = g, column = 5).value
    if regulation is not None:
        regulation = regulation.strip()
    # Initialize if necessary
    if team1 not in record:
        record[team1] = [0, 0, 0]
        # Always start with 100 as every team's rating (using iteration to solve the recursive problem)
        ratings[team1] = 100
    if team2 not in record:
        record[team2] = [0, 0, 0]
        # Always start with 100 as every team's rating (using iteration to solve the recursive problem)
        ratings[team2] = 100
    
    # Check to see winner
    # Counts all OT as ties
    if (regulation != '' and regulation is not None) or team1score == team2score:
        record[team1][2] += 1
        record[team2][2] += 1        
    elif team1score > team2score:
        record[team1][0] += 1
        record[team2][1] += 1
    else:
        record[team1][1] += 1
        record[team2][0] += 1
    
    # use 55/45 weights for OT
    #if team1score == team2score:
        #record[team1][2] += 1
        #record[team2][2] += 1        
    #elif team1score > team2score and (regulation == '' or regulation is None):
        #record[team1][0] += 1
        #record[team2][1] += 1
    #elif team1score < team2score and (regulation == '' or regulation is None):
        #record[team1][1] += 1
        #record[team2][0] += 1
    #elif team1score > team2score:
        #record[team1][0] += 0.55
        #record[team1][1] += 0.45
        #record[team2][1] += 0.55
        #record[team2][0] += 0.45
    #else:
        #record[team1][1] += 0.55
        #record[team1][0] += 0.45
        #record[team2][0] += 0.55
        #record[team2][1] += 0.45
    
    games.append((team1, team2))
    # Increment counter for reading the spreadsheet
    g += 1
    
# Recursion for getting accurate ratings
iterations = 0
while not done:
    # Initialize the flag to True each time
    flag = True
    
    # Clear expected wins each iteration
    expected.clear()
    # For every game, calculate:
    for game in games:
        # Weighting factor (1 divided by the sum of the ratings of the 2 teams)
        wf = 1 / (ratings[game[0]] + ratings[game[1]])
        # Check to see if each team is in expected wins dictionary
        if game[0] in expected:
            # Multiply team's rating by weighting factor and add to sum
            expected[game[0]] += (ratings[game[0]] * wf)
        else:
            # Multiply team's rating by weighting factor and initialize as expected wins
            expected[game[0]] = (ratings[game[0]] * wf)
        if game[1] in expected:
            # Multiply team's rating by weighting factor and add to sum
            expected[game[1]] += (ratings[game[1]] * wf)
        else:
            # Multiply team's rating by weighting factor and initialize as expected wins
            expected[game[1]] = (ratings[game[1]] * wf)
    
    # For every team, calculate:
    for key in ratings:
        wins = record[key][0] + 0.5 * record[key][2]
        if key not in info:
            # If team is undefeated, we'll get a divide by 0 error, so we set the ratio to 25
            # Note: this formula is most effective when there are no undefeated or winless teams
            # and a chain of wins (or ties) can be made from every team to any other team
            if record[key][1] == 0 and record[key][2] == 0:
                ratio = 25
            # If the team is winless, everything would automatically be 0, so set the ratio to 1/25
            elif record[key][0] == 0 and record[key][2] == 0:
                ratio = 1/25
            # Win ratio = wins / losses
            else:
                ratio = wins / (record[key][1] + 0.5 * record[key][2])
            # Add the team's win %, win ratio and sos is intialized to 0 (to be calculated later)
            teaminfo = [wins / (record[key][0] + record[key][1] + record[key][2]), ratio, 0]
            # Add the list of team info to the info dict
            info[key] = teaminfo
        # New rating for the team equals the team's wins divided by expected wins multiplied by the old rating
        new_ratings[key] = (wins / expected[key]) * ratings[key]
        # Update the SOS for the team
        info[key][2] = new_ratings[key] / info[key][1]
        # If the difference between old rating and new rating <= DELTA
        # If flag is true, that means so far every team's new rating has been within DELTA
        # (since flag is initialized to true)
        if abs(ratings[key] - new_ratings[key]) <= DELTA and abs(wins - expected[key]) <= DELTA and flag:
            done = True
        # If the difference is greater than DELTA, we must continue the recursion
        # If flag is false, one team has already failed the DELTA test and we must continue the recursion
        else:
            flag = False
            done = False
    # After going through all teams, update ratings
    ratings = new_ratings.copy()
    iterations += 1

# Scale the ratings to an average of 100
for i in range(10):
    scale_wins = 0
    for key in ratings:
        scale_wins += 100 / (100 + ratings[key])
    # Divide by half the total number of teams
    scale = scale_wins / (len(ratings) / 2)
    
    # Adjust every team's rating and SOS according to scale
    for key in ratings:
        ratings[key] *= scale
        info[key][2] = ratings[key] / info[key][1]

# Sort the teams by their rating
sortedratings = sorted(ratings.items(), key=lambda kv: kv[1], reverse = True)
# Write to original file
wb.create_sheet('Output')
# For outputting the results into the spreadsheet
out = wb['Output']
for r in range(len(ratings) + 1):
    # Header row
    if r == 0:
        out.cell(row = r + 1, column = 1).value = 'Rank'
        out.cell(row = r + 1, column = 2).value = 'Team'
        out.cell(row = r + 1, column = 3).value = 'Rating'
        out.cell(row = r + 1, column = 4).value = 'Win % Rank'
        out.cell(row = r + 1, column = 5).value = 'Record'
        out.cell(row = r + 1, column = 6).value = 'Win %'
        out.cell(row = r + 1, column = 7).value = 'Win Ratio'
        out.cell(row = r + 1, column = 8).value = 'SOS Rank'
        out.cell(row = r + 1, column = 9).value = 'SOS'
    else:
        # Nine columns
        for col in range(9):
            # Col 0 is the rank
            if col == 0:
                out.cell(row = r + 1, column = col + 1).value = r
            # Col 1 is the team name, col 2 is the rating
            elif col < 3:
                out.cell(row = r + 1, column = col + 1).value = sortedratings[r - 1][col - 1]
            # Formula to calculate the rank of winning %
            elif col == 3:
                out.cell(row = r + 1, column = col + 1).value = "=RANK(F" + str(r + 1) + \
                    ",F2:F" + str(len(sortedratings) + 1) + ")"
            # Output record
            elif col == 4:
                out.cell(row = r + 1, column = col + 1).value = str(record[sortedratings[r - 1][0]][0]) \
                    + "-" + str(record[sortedratings[r - 1][0]][1]) + "-" + str(record[sortedratings[r - 1][0]][2])
            # Output win % (col 5) and win ratio (col 6)
            elif col < 7:
                out.cell(row = r + 1, column = col + 1).value = info[sortedratings[r - 1][0]][col - 5]
            # Formula to calculate the SOS rank
            elif col == 7:
                out.cell(row = r + 1, column = col + 1).value = "=RANK(I" + str(r + 1) + \
                    ",I2:I" + str(len(sortedratings) + 1) + ")"
            # Output SOS (col 8)
            else:
                out.cell(row = r + 1, column = col + 1).value = info[sortedratings[r - 1][0]][col - 6]

# Save the sheet with the output
wb.save('Bradley-Terry Spreadsheet NCAA Hockey.xlsx')
