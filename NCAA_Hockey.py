import openpyxl

def writeOutput(teamstats, filename):
    # Make ratings dict for output
    ratings = dict()
    for t in teamstats:
        ratings[t] = teamstats[t]["rating"]

    # Sort the teams by their rating
    sortedratings = sorted(ratings.items(), key=lambda kv: kv[1], reverse = True)

    # Create new sheet for output
    wb = openpyxl.Workbook()
    wb.create_sheet('Output')
    # For outputting the results into the spreadsheet
    out = wb['Output']
    
    for r in range(len(sortedratings) + 1):
        # Header row
        if r == 0:
            out.cell(row = r + 1, column = 1).value = 'Rank'
            out.cell(row = r + 1, column = 2).value = 'Team'
            out.cell(row = r + 1, column = 3).value = 'Rating'
            out.cell(row = r + 1, column = 4).value = 'Win % Rank'
            out.cell(row = r + 1, column = 5).value = 'Record'
            out.cell(row = r + 1, column = 6).value = 'Win %'
            out.cell(row = r + 1, column = 7).value = 'Win Ratio'
            out.cell(row = r + 1, column = 8).value = 'SOS Rank'
            out.cell(row = r + 1, column = 9).value = 'SOS'
        else:
            # Nine columns
            for col in range(9):
                # Col 0 is the rank
                if col == 0:
                    out.cell(row = r + 1, column = col + 1).value = r
                # Col 1 is the team name, col 2 is the rating
                elif col < 3:
                    out.cell(row = r + 1, column = col + 1).value = sortedratings[r - 1][col - 1]
                # Formula to calculate the rank of winning %
                elif col == 3:
                    out.cell(row = r + 1, column = col + 1).value = "=RANK(F" + str(r + 1) + \
                        ",F2:F" + str(len(sortedratings) + 1) + ")"
                # Output record
                elif col == 4:
                    out.cell(row = r + 1, column = col + 1).value = str(teamstats[sortedratings[r - 1][0]]["record"][0]) \
                        + "-" + str(teamstats[sortedratings[r - 1][0]]["record"][1]) + "-" \
                        + str(teamstats[sortedratings[r - 1][0]]["record"][2])
                # Output win % (col 5) and win ratio (col 6)
                elif col < 7:
                    out.cell(row = r + 1, column = col + 1).value = teamstats[sortedratings[r - 1][0]]["info"][col - 5]
                # Formula to calculate the SOS rank
                elif col == 7:
                    out.cell(row = r + 1, column = col + 1).value = "=RANK(I" + str(r + 1) + \
                        ",I2:I" + str(len(sortedratings) + 1) + ")"
                # Output SOS (col 8)
                else:
                    out.cell(row = r + 1, column = col + 1).value = teamstats[sortedratings[r - 1][0]]["info"][col - 6]

    # Save the sheet with the output
    wb.save(filename)

def scaleRatings(teamstats):
    # Scale the ratings to an average of 100
    for _ in range(10):
        scale_wins = 0
        for key in teamstats:
            scale_wins += 100 / (100 + teamstats[key]["rating"])
        # Divide by half the total number of teams
        scale = scale_wins / (len(teamstats) / 2)
        
        # Adjust every team's rating and SOS according to scale
        for key in teamstats:
            teamstats[key]["rating"] *= scale
            teamstats[key]["info"][2] = teamstats[key]["rating"] / teamstats[key]["info"][1]

def calculateRatings(teamstats, games):
    # Value to compare every rating to and then do recursion
    DELTA = 0.0001
    # Main boolean used with while loop for recursion based on flag
    DONE = False
    # If all values are within DELTA (i.e. should recursion finish)
    FLAG = True

    # Recursion for getting accurate ratings
    while not DONE:
        # Initialize the flag to True each time
        FLAG = True
        
        # Clear expected wins each iteration
        for t in teamstats:
            teamstats[t]["expected"] = 0
        
        # For every game, calculate:
        for game in games:
            # Weighting factor (1 divided by the sum of the ratings of the 2 teams)
            wf = 1 / (teamstats[game[0]]["rating"] + teamstats[game[1]]["rating"])
            # Multiply team's rating by weighting factor and add to sum
            teamstats[game[0]]["expected"] += (teamstats[game[0]]["rating"] * wf)
            # Multiply team's rating by weighting factor and add to sum
            teamstats[game[1]]["expected"] += (teamstats[game[1]]["rating"] * wf)
        
        # For every team, calculate:
        for key in teamstats:
            # New rating for the team equals the team's wins divided by expected wins multiplied by the old rating
            wins = teamstats[key]["record"][0] + 0.5 * teamstats[key]["record"][2]
            teamstats[key]["new_rating"] = (wins / teamstats[key]["expected"]) * teamstats[key]["rating"]
            
            # Update the SOS for the team
            teamstats[key]["info"][2] = teamstats[key]["new_rating"] / teamstats[key]["info"][1]
            
            # If the difference between old rating and new rating <= DELTA
            # If flag is true, that means so far every team's new rating has been within DELTA
            # (since flag is initialized to true)
            if abs(teamstats[key]["rating"] - teamstats[key]["new_rating"]) <= DELTA \
            and abs(wins - teamstats[key]["expected"]) <= DELTA and FLAG:
                DONE = True
            # If the difference is greater than DELTA, we must continue the recursion
            # If flag is false, one team has already failed the DELTA test and we must continue the recursion
            else:
                FLAG = False
                DONE = False
        
        # After going through all teams, update ratings
        for t in teamstats:
            teamstats[t]["rating"] = teamstats[t]["new_rating"]

def updateInfo(teamstats):
    # Update info after every game is read
    for key in teamstats:
        # If team is undefeated or winless, we'll get a divide by 0 error
        # Note: this formula is most effective when there are no undefeated or winless teams
        # and a chain of wins (or ties) can be made from every team to any other team
        wins = teamstats[key]["record"][0] + 0.5 * teamstats[key]["record"][2]
        ratio = wins / (teamstats[key]["record"][1] + 0.5 * teamstats[key]["record"][2])
        # Add the team's win %, win ratio and sos is intialized to 0 (to be summed later)
        teaminfo = [wins / (teamstats[key]["record"][0] + teamstats[key]["record"][1] + teamstats[key]["record"][2]), ratio, 0]
        # Add the list of team info to the info dict
        teamstats[key]["info"] = teaminfo

def readGames(teamstats, games, filename):
    # Open the spreadsheet and assign the first 2 sheets
    # Replace with your xlsx file name
    wb = openpyxl.load_workbook(filename) 
    firstsheet = wb.worksheets[0]

    # Read in every game with the teams
    g = 2
    while g < firstsheet.max_row + 1:
        # Skip exhibition games
        marker = firstsheet.cell(row = g, column = 7).value.strip()
        if marker == 'ex' or marker == 'n3':
            g += 1
            continue

        team1 = firstsheet.cell(row = g, column = 1).value.strip()
        team1score = firstsheet.cell(row = g, column = 2).value
        team2 = firstsheet.cell(row = g, column = 4).value.strip()
        team2score = firstsheet.cell(row = g, column = 5).value
        regulation = firstsheet.cell(row = g, column = 6).value
        # Remove possible padding
        if regulation is not None:
            regulation = regulation.strip()
        # Game hasn't been played yet
        if team1score == '' or team1score is None or team2score == '' or team2score is None:
            g += 1
            continue
        
        # Initialize if necessary
        if team1 not in teamstats:
            # Initialize team dict
            teamstats[team1] = dict()
            # Initialize record to 0-0
            teamstats[team1]["record"] = [0, 0, 0]
            # Always start with 100 as every team's rating (using iteration to solve the recursive problem)
            teamstats[team1]["rating"] = 100
        if team2 not in teamstats:
            # Initialize team dict
            teamstats[team2] = dict()
            # Initialize record to 0-0
            teamstats[team2]["record"] = [0, 0, 0]
            # Always start with 100 as every team's rating (using iteration to solve the recursive problem)
            teamstats[team2]["rating"] = 100
        
        # Check to see winner
        # Counts all OT as ties
        if (regulation != '' and regulation is not None) or team1score == team2score:
            teamstats[team1]["record"][2] += 1
            teamstats[team2]["record"][2] += 1
        elif team1score > team2score:
            teamstats[team1]["record"][0] += 1
            teamstats[team2]["record"][1] += 1
        else:
            teamstats[team1]["record"][1] += 1
            teamstats[team2]["record"][0] += 1
        
        # use 55/45 weights for OT
        #if team1score == team2score:
            # teamstats[team1]["record"][2] += 1
            # teamstats[team2]["record"][2] += 1
        #elif team1score > team2score and (regulation == '' or regulation is None):
            # teamstats[team1]["record"][0] += 1
            # teamstats[team2]["record"][1] += 1
        #elif team1score < team2score and (regulation == '' or regulation is None):
            # teamstats[team1]["record"][1] += 1
            # teamstats[team2]["record"][0] += 1
        #elif team1score > team2score:
            # teamstats[team1]["record"][0] += 0.55
            # teamstats[team1]["record"][1] += 0.45
            # teamstats[team2]["record"][1] += 0.55
            # teamstats[team2]["record"][0] += 0.45
        #else:
            # teamstats[team1]["record"][1] += 0.55
            # teamstats[team1]["record"][0] += 0.45
            # teamstats[team2]["record"][0] += 0.55
            # teamstats[team2]["record"][1] += 0.45
        
        games.append((team1, team2))
        # Increment counter for reading the spreadsheet
        g += 1

def main(read):
    # List of tuples which are the games; each tuple is (winner, loser)
    games = []

    # Mega dictionary with all info
    # key = team name
    # value = dictionary with: rating (int), record (list of [wins, losses]), expected wins (int)
    # new rating (int) and info (list of [win %, win ratio, strength of schedule])
    teamstats = dict()

    readGames(teamstats, games, read)
    updateInfo(teamstats)
    calculateRatings(teamstats, games)
    scaleRatings(teamstats)

    return teamstats

if __name__ == "__main__":
    # Replace with your filenames
    teamstats = main('../women.xlsx')
    writeOutput(teamstats, 'Bradley-Terry Spreadsheet NCAA Women Hockey.xlsx')
